from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPORT_DIR = Path(r"C:\Users\gmhome\Downloads\Availibility report")
PROJECT_DIR = Path(r"C:\Users\gmhome\SynologyDrive\Ragmaterials\exam\QNA-Generator")

HEADER_FILL = PatternFill("solid", fgColor="001F4E78")
HEADER_FONT = Font(color="00FFFFFF", bold=True)
COMPLETE_FILL = PatternFill("solid", fgColor="00E2F0D9")
MISSING_MS_FILL = PatternFill("solid", fgColor="00FFF2CC")
MISSING_QP_FILL = PatternFill("solid", fgColor="00FCE4D6")
MISSING_BOTH_FILL = PatternFill("solid", fgColor="00F4CCCC")
SUMMARY_GOOD_FILL = PatternFill("solid", fgColor="00E2F0D9")
SUMMARY_MID_FILL = PatternFill("solid", fgColor="00FFF2CC")
SUMMARY_LOW_FILL = PatternFill("solid", fgColor="00FCE4D6")
SUMMARY_BAD_FILL = PatternFill("solid", fgColor="00F4CCCC")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
SESSION_ORDER = {"M": 0, "S": 1, "W": 2, "SPEC": 3, "?": 4}
BOARD_ORDER = {"CIE": 0, "Cambridge": 0, "Edexcel": 1, "AQA": 2, "SL": 0, "HL": 1}


def load_rows(csv_name: str) -> list[dict[str, str]]:
    path = REPORT_DIR / csv_name
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def infer_igcse_board(row: dict[str, str]) -> str:
    text = " ".join(row.get(key, "") for key in ("Paper", "QP", "MS", "ER", "IN", "Transcript")).lower()
    if "edexcel" in text or "pearson" in text:
        return "Edexcel"
    if "cambridge" in text or "cie" in text:
        return "CIE"
    return "Unknown"


def infer_alevel_board(row: dict[str, str]) -> str:
    syllabus = row.get("Syllabus", "").strip()
    if syllabus:
        normalized = syllabus.lower()
        if normalized == "edexcel":
            return "Edexcel"
        if normalized == "aqa":
            return "AQA"
        if normalized in {"cambridge", "cie"}:
            return "CIE"
    text = " ".join(row.get(key, "") for key in ("Paper", "QP", "MS", "ER", "IN", "Transcript")).lower()
    if "edexcel" in text or "pearson" in text:
        return "Edexcel"
    if "aqa" in text:
        return "AQA"
    if "cambridge" in text or "cie" in text:
        return "CIE"
    return "Unknown"


def infer_ib_level(row: dict[str, str]) -> str:
    text = " ".join(row.get(key, "") for key in ("Paper", "QP", "MS", "ER", "IN", "Transcript")).lower()
    if "-hl-" in text or "\\hl\\" in text or "/hl/" in text:
        return "HL"
    if "-sl-" in text or "\\sl\\" in text or "/sl/" in text:
        return "SL"
    return "Unknown"


def status_fill(status: str) -> PatternFill:
    if status == "Complete":
        return COMPLETE_FILL
    if status == "Missing MS":
        return MISSING_MS_FILL
    if status == "Missing QP":
        return MISSING_QP_FILL
    return MISSING_BOTH_FILL


def summary_fill(complete_ratio: float) -> PatternFill:
    if complete_ratio >= 0.95:
        return SUMMARY_GOOD_FILL
    if complete_ratio >= 0.75:
        return SUMMARY_MID_FILL
    if complete_ratio >= 0.5:
        return SUMMARY_LOW_FILL
    return SUMMARY_BAD_FILL


def normalize_missing_text(status: str) -> str | None:
    if status == "Missing QP":
        return "QP"
    if status == "Missing MS":
        return "MS"
    if status == "No QP/MS":
        return "QP + MS"
    return None


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = name[:31]
    if base not in used:
        used.add(base)
        return base
    index = 2
    while True:
        suffix = f" ({index})"
        candidate = f"{name[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        index += 1


def build_split_workbook(
    rows: list[dict[str, str]],
    qualifier_name: str,
    bucket_label: str,
    bucket_fn,
    output_name: str,
) -> None:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        subject = row.get("Subject", "").strip()
        if not subject:
            continue
        bucket = bucket_fn(row)
        grouped[(subject, bucket)].append(row)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append([f"Subject / {bucket_label}", "Total", "Complete", "Missing QP", "Missing MS", "Missing Both"])
    for cell in summary_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    used_sheet_names = {"Summary"}
    ordered_keys = sorted(
        grouped,
        key=lambda item: (
            item[0].lower(),
            BOARD_ORDER.get(item[1], 99),
            item[1].lower(),
        ),
    )

    for subject, bucket in ordered_keys:
        sheet_title = safe_sheet_name(f"{subject} - {bucket}", used_sheet_names)
        ws = wb.create_sheet(sheet_title)
        ws.append(["Year", "Session", "Code", "Paper", "Status", "Missing"])
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        subject_rows = sorted(
            grouped[(subject, bucket)],
            key=lambda row: (
                int(row.get("Year", "9999")) if row.get("Year", "").isdigit() else 9999,
                SESSION_ORDER.get(row.get("Session", "").strip(), 9),
                row.get("Code", ""),
                row.get("Paper", ""),
            ),
        )

        complete = missing_qp = missing_ms = missing_both = 0
        for row in subject_rows:
            status = row.get("Status", "").strip()
            if status == "Complete":
                complete += 1
            elif status == "Missing QP":
                missing_qp += 1
            elif status == "Missing MS":
                missing_ms += 1
            else:
                missing_both += 1

            ws.append(
                [
                    row.get("Year", ""),
                    row.get("Session", ""),
                    row.get("Code", "") or None,
                    row.get("Paper", ""),
                    status,
                    normalize_missing_text(status),
                ]
            )
            row_idx = ws.max_row
            fill = status_fill(status)
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.alignment = CENTER if cell.column != 4 else LEFT

        for col, width in {"A": 10, "B": 10, "C": 12, "D": 50, "E": 16, "F": 14}.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"

        total = len(subject_rows)
        summary_ws.append([f"{subject} - {bucket}", total, complete, missing_qp, missing_ms, missing_both])
        ratio = complete / total if total else 0
        for cell in summary_ws[summary_ws.max_row]:
            cell.fill = summary_fill(ratio)
            cell.alignment = CENTER if cell.column != 1 else LEFT

    for col, width in {"A": 36, "B": 10, "C": 10, "D": 12, "E": 12, "F": 14}.items():
        summary_ws.column_dimensions[col].width = width
    summary_ws.freeze_panes = "A2"

    project_out = PROJECT_DIR / output_name
    report_out = REPORT_DIR / output_name
    wb.save(project_out)
    wb.save(report_out)
    print(f"Saved {project_out}")
    print(f"Saved {report_out}")


def main() -> None:
    build_split_workbook(
        load_rows("IGCSE.csv"),
        "IGCSE",
        "Board",
        infer_igcse_board,
        "IGCSE_by_subject_updated.xlsx",
    )
    build_split_workbook(
        load_rows("A-LEVEL.csv"),
        "A-Level",
        "Board",
        infer_alevel_board,
        "A-LEVEL_by_subject_updated.xlsx",
    )
    build_split_workbook(
        load_rows("IB.csv"),
        "IB",
        "Level",
        infer_ib_level,
        "IB_by_subject_updated.xlsx",
    )


if __name__ == "__main__":
    main()
